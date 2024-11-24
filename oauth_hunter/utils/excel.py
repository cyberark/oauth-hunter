import openpyxl
from openpyxl.styles import PatternFill, Font
from engine.scenarios import BASE_HEADERS, HEADERS, REDIRECT_URI_SCENARIOS
from utils.common_utils import get_domain_name
import os

from . import config

def get_unique_filename(filename):
    """
    Check if the filename already exists. If it does, append a number to the filename
    and keep checking until a unique filename is found.
    """
    base, extension = os.path.splitext(filename)
    counter = 1
    new_filename = filename

    file_path = f"{config.EXCEL_FOLDER}/{new_filename}"
    # Check if the file exists, if it does, keep increasing the number
    while os.path.exists(file_path):
        new_filename = f"{base}_{counter}{extension}"
        file_path = f"{config.EXCEL_FOLDER}/{new_filename}"
        counter += 1

    return new_filename


def handle_excel_file(create_excel_arg, overwrite_flag):
    """
    Handles the creation or overwriting of an Excel file based on command-line arguments.

    Args:
        create_excel_arg (str or None): The filename provided with --create-excel or None.
        overwrite_flag (bool): Whether --overwrite is specified.

    Returns:
        str: The full path to the Excel file.

    """
    if create_excel_arg:
        # User specified a filename or default filename
        excel_filename = create_excel_arg
        excel_full_path = f"{config.EXCEL_FOLDER}/{excel_filename}"

        if overwrite_flag:
            # If --overwrite is specified
            if os.path.exists(excel_full_path):
                # Set the path to the existing file
                print(f"Using existing file for overwrite: {excel_full_path}")
            else:
                # File does not exist, handle as needed
                print(f"File not found for overwrite: {excel_full_path}")
                create_excel_file(excel_full_path)

        else:
            # If --overwrite is not specified
            if os.path.exists(excel_full_path):
                # Generate a unique filename to avoid overwriting
                unique_filename = get_unique_filename(excel_filename)
                excel_full_path = f"results/{unique_filename}"
                create_excel_file(excel_full_path)
            else:
                # Create the file if it does not exist
                create_excel_file(excel_full_path)

    else:
        # No --create-excel argument provided, use the default filename
        excel_full_path = f"{config.EXCEL_FOLDER}/{config.EXCEL_DEFAULT_FILENAME}"

        if overwrite_flag:
            # If --overwrite is specified
            if os.path.exists(excel_full_path):
                # Set the path to the existing file
                print(f"Using existing default file for overwrite: {excel_full_path}")
            else:
                # File does not exist, handle as needed
                print(f"Default file not found for overwrite: {excel_full_path}")
                create_excel_file(excel_full_path)

        else:
            # Create the file if it does not exist
            if not os.path.exists(excel_full_path):
                create_excel_file(excel_full_path)

    # Update the full path in config

    config.EXCEL_FULL_PATH = excel_full_path
    #return excel_full_path

def create_excel_file(filename):
    #filename = get_unique_filename(filename)
    directory = os.path.dirname(filename)
    if not os.path.exists(directory):
        os.makedirs(directory)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OAuth Test Results"

    # Header cells with color and bold font
    header_fill = PatternFill(start_color="b6d7a8", end_color="b6d7a8", fill_type="solid")
    header_font = Font(bold=True)
    redirect_uri_fill = PatternFill(start_color="ea9999", end_color="ea9999", fill_type="solid")
    state_fill = PatternFill(start_color="ffe599", end_color="ffe599", fill_type="solid")

    # Total number of redirect_uri scenarios
    num_redirect_uri_scenarios = len([scenario for scenario in REDIRECT_URI_SCENARIOS])

   # ws.append(["", "", "", "", "", "", ""] + [""] * (len(HEADERS) - 10))
    ws.append([""] * len(HEADERS))
    ws.append(HEADERS)  # Adding the main headers

    for col_num, header in enumerate(HEADERS, 1):  # Adjusting for the additional header row
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    # Redirect URI header configuration
    redirect_uri_index = len(BASE_HEADERS) + 1
    state_index = redirect_uri_index + num_redirect_uri_scenarios

    # Setting color for the "redirect_uri" header cells
    for col_num in range(redirect_uri_index, state_index):  # Columns C to E (1-based indexing)
        cell = ws.cell(row=1, column=col_num)
        cell.fill = redirect_uri_fill
        cell.value = ""  # Empty value for the merged cell

    # Setting color for the "State" header cells
    for col_num in range(state_index, state_index + 3):  # Columns F to G (1-based indexing)
        cell = ws.cell(row=1, column=col_num)
        cell.fill = state_fill
        cell.value = ""  # Empty value for the merged cell

    # Set the "redirect_uri" cell to the appropriate value
    ws.cell(row=1, column=redirect_uri_index).value = "redirect_uri"
    ws.merge_cells(start_row=1, start_column=redirect_uri_index, end_row=1, end_column=state_index - 1)

    # Set the "State" cell to the appropriate value
    ws.cell(row=1, column=state_index).value = "State"
    ws.merge_cells(start_row=1, start_column=state_index, end_row=1, end_column=state_index + 2)

    # Freeze the rows under the headers
    ws.freeze_panes = "A3"
    wb.save(filename)

def add_test_result_to_excel(filename, test_result):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    green_fill = PatternFill(start_color="b7e1cd", end_color="b7e1cd", fill_type="solid")
    red_fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid")

    def get_fill(value, true_color, false_color):
        return true_color if value else false_color

    row = [
        get_domain_name(test_result.domain),
        test_result.domain,
        test_result.oauth_type,
        test_result.scope,
        test_result.response_type,
        test_result.client_id,
        test_result.original_url
    ]

    # go over the OAuth redirect scenarios and add it to the row
    for scenario in test_result.test_scenarios:
        result = "yes" if scenario.is_succeed else "no"
        row.append(result)

    ws.append(row)
    row_num = ws.max_row
    # ws.cell(row=row_num, column=4).fill = get_fill(test_result.is_redirect_uri_subdomain_verified, green_fill, red_fill)
    # ws.cell(row=row_num, column=5).fill = get_fill(test_result.is_redirect_uri_any_path_verified, green_fill, red_fill)
    # ws.cell(row=row_num, column=6).fill = get_fill(test_result.is_redirect_uri_verified, green_fill, red_fill)
    # ws.cell(row=row_num, column=7).fill = get_fill(test_result.is_state_verified, green_fill, red_fill)
    # ws.cell(row=row_num, column=8).fill = get_fill(test_result.can_omit_state, red_fill, green_fill)

    # Fill the corresponding cells based on the scenario results
    for col_num, scenario in enumerate(test_result.test_scenarios, start=len(HEADERS) - len(REDIRECT_URI_SCENARIOS) + 1):
        fill_color = green_fill if scenario.is_succeed else red_fill
        ws.cell(row=row_num, column=col_num).fill = fill_color

    wb.save(filename)

